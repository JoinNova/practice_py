import pandas as pd
import openpyxl
'''
data=pd.read_csv('/home/nova/bigdata/python/today/crime_in_seoul.csv')
data= pd.DataFrame(data)
print(data)
print('='*80)
data.to_excel('police_test01.xlsx', sheet_name='sheet1', index=False)
result=data.sort_values(['관서명'], ascending=True)
print(result)
result.to_excel('police_test02.xlsx', sheet_name='sheet1', index=False)
'''
'''
from re import sub
result=str()
chk=0 

data=open('/home/nova/bigdata/python/today/crime_in_seoul.csv','r')
for _ in data:
    _=sub(' ','',_)
    if _.isalpha()==0:
        s=_.strip()
        for i in range(4):
            s=sub(',1,',',0,',s)
            print(s)
        s=sub(r'"2{1},\d{3}"','"2,000"',s)
        s=sub(r'"3{1},\d{3}"','"2,000"',s)
        s=sub(r'"4{1},\d{3}"','"2,000"',s)
        #print(s)
        result+=s
        result+='\n'
    else:
        result+=_
           
print(result)
#print('='*80)
data.close()                                 ##____raw데이터파일 종료
  
file_ = open('/home/nova/bigdata/python/today/police_revert.xlsx', 'w')    ##____정제데이터저장할 파일 오픈
file_.write(result)                                               ##____정제데이터 파일에 쓰기
file_.close()  
'''
'''
from re import sub
import pandas as pd
import numpy as np
from openpyxl import load_workbook

data = load_workbook(filename = 'test.xlsx')
print(data)
    

#data.close()                                 ##____raw데이터파일 종료
'''
'''
from re import sub                    ###문자소거시 사용하는 re클래스호출
from openpyxl import load_workbook    ###엑셀파일 열어주는 openpyxl 호출
wb = load_workbook(filename='test.xlsx', read_only=True)  #엑셀파일명
ws = wb['sheet1']                   ###엑셀시트명
result=str()                        ###내가 새로만들 엑셀 파일 변수생성
for row in ws.rows:                 ###시트에서 1행씩 불러오기
    for cell in row:                ### 한 행에서 한셀씩 소환
        #chk=0
        #print(type(cell.value[0]))
        try:                        ###문자와 숫자를 구별하기위한 트릭<억지로 오류내기>
            if cell.value[0].isdigit()==True:     ###문자 또는  오류나서 except로 
                if int(sub(',','',cell.value))>=2000:    ###2000보다 클경우 "2,000",으로 저장 
                    result+='"2,000",'
                elif int(sub(',','',cell.value))>=1000:   ###1000보다 클경우  "값",로 저장
                    #print(str(cell.value))
                    result+='"'+str(cell.value)+'",'
                else:
                    #print(str(cell.value))
                    result+=str(cell.value)+','           ###1000이하일경우 "값",로 저장
            else:
                result+=sub(' ','',str(cell.value))+',' ###문자일경우 문자사이 공백 제거 하고 저장
        except:
            if cell.value==1 :           ###숫자가 한자리일경우 cell.value[0]가 존재하지않으므로 except
                #print(cell.value)
                result+='0,'            ### 0,으로 변환
            else:
                result+=sub(' ','',str(cell.value))+','  ###문자 또는 숫자 한자리일경우 value, 로 저장
    result+='\n'             ###각 행 마다 개행문자(\n) 추가.
        #if cell.value==1:
            #print(cell.value)
        #result+=_
print(result)        ###새로만든 엑셀파일 결과 확인 

file_ = open('/home/nova/bigdata/python/today/revert_openpyxl.xlsx', 'w')    ##____정제데이터저장할 파일 오픈
file_.write(result)                                               ##____정제데이터 파일에 쓰기
file_.close() 
'''
'''
import csv
import pip
import matplotlib.pyplot
import pandas as pd
'''
'''
file_path = '/home/nova/bigdata/python/today/gong.csv'
df = open(file_path)
result=str()
chk=0
for _ in df:
    chk+=1
    if chk==1:
        result+=_
    if _.count('공주시')!=0 and _.count('음식')!=0:
        #print(_[:-1])
        result+=_
file_ = open('/home/nova/bigdata/python/today/gong_cv.xlsx', 'w')    ##____정제데이터저장할 파일 오픈
file_.write(result)                                               ##____정제데이터 파일에 쓰기
file_.close() 
'''
'''
from pandas import DataFrame, Series
 
df = pd.read_csv(r'/home/nova/bigdata/python/today/gong.csv', engine='python',
                 encoding='UTF-8')
 
a = df.loc[df['시군구명'] == '공주시', :] 
a1 = a.loc[a['상권업종대분류명'] == '음식', :]
 
a1.to_excel('gong_cv2.xlsx', index=False)
'''

'''
#df.head(10)
#df.info()
df.drop(['상가업소번호','지점명','시도명','상권업종대분류코드','상권업종중분류코드',
         '상권업종중분류명','상권업종소분류코드','상권업종소분류명','표준산업분류코드',
         '표준산업분류명','시도코드','시군구코드','행정동코드',
         '행정동명','법정동코드','법정동명','지번코드','대지구분코드',
         '대지구분명','지번본번지','건물부번지','지번부번지','지번주소','도로명코드',
         '도로명','건물본번지','건물관리번호','건물명','도로명주소','구우편번호',
         '신우편번호','동정보','층정보','호정보','위도','경도'],axis=1,inplace=True)
df.columns.values
df.to_csv("gong_cv01.csv")
'''
